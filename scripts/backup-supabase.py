#!/usr/bin/env python3
"""
Backup Supabase InternLog → Excel → SFTP OVH
Exporte toutes les tables principales en un fichier Excel multi-onglets
puis l'upload sur l'hébergement OVH KGN dans /backups/internlog/

=== Configuration (obligatoire) ===

Créer un fichier scripts/.env.scripts (gitignored) avec :

    SUPABASE_URL=https://xxxxxxxx.supabase.co
    SUPABASE_SERVICE_ROLE_KEY=eyJ...
    OVH_HOST=ssh.clusterXXX.hosting.ovh.net
    OVH_USER=xxxxxx
    OVH_PASS=xxxxxx
    OVH_REMOTE_DIR=/home/xxxxxx/backups/internlog

Les variables peuvent aussi être définies directement dans l'environnement
shell. Le script refuse de démarrer si une variable obligatoire est absente.
"""

import os
import sys
import json
import subprocess
import tempfile
import ssl
from datetime import datetime
from pathlib import Path
from urllib.request import Request, urlopen

# Fix certificats SSL macOS
ssl._create_default_https_context = ssl._create_unverified_context


def load_env_file(path: Path) -> None:
    """Charge un fichier .env simple (KEY=VALUE) dans os.environ sans écraser."""
    if not path.is_file():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        os.environ.setdefault(key, value)


def require_env(name: str) -> str:
    """Récupère une variable d'environnement ou sort en erreur."""
    value = os.environ.get(name)
    if not value:
        print(
            f"✗ Variable d'environnement manquante : {name}\n"
            f"  → Définir dans scripts/.env.scripts ou dans le shell.",
            file=sys.stderr,
        )
        sys.exit(1)
    return value


# --- Chargement config ---
SCRIPT_DIR = Path(__file__).resolve().parent
load_env_file(SCRIPT_DIR / ".env.scripts")

SUPABASE_URL = require_env("SUPABASE_URL")
SUPABASE_KEY = require_env("SUPABASE_SERVICE_ROLE_KEY")
OVH_HOST = require_env("OVH_HOST")
OVH_USER = require_env("OVH_USER")
OVH_PASS = require_env("OVH_PASS")
OVH_REMOTE_DIR = os.environ.get("OVH_REMOTE_DIR", f"/home/{OVH_USER}/backups/internlog")

TABLES = [
    "profiles",
    "des_registry",
    "hospitals",
    "specialties",
    "procedures",
    "entries",
    "gardes",
    "des_objectives",
    "subscriptions",
    "institutional_seats",
    "seat_assignments",
    "adhesion_requests",
    "patient_followups",
    "supervisor_assignments",
    "audit_log",
]


def fetch_table(table_name: str) -> list[dict]:
    """Récupère toutes les lignes d'une table Supabase via REST API."""
    url = f"{SUPABASE_URL}/rest/v1/{table_name}?select=*&limit=10000"
    req = Request(url, headers={
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
    })
    try:
        with urlopen(req) as resp:
            return json.loads(resp.read().decode())
    except Exception as e:
        print(f"  ⚠ Table '{table_name}' : {e}")
        return []


def create_excel(tables_data: dict[str, list[dict]], output_path: str):
    """Crée un fichier Excel avec un onglet par table."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="0891B2", end_color="0891B2", fill_type="solid")

    for table_name, rows in tables_data.items():
        if not rows:
            ws = wb.create_sheet(title=table_name[:31])
            ws.append(["Aucune donnée"])
            continue

        ws = wb.create_sheet(title=table_name[:31])
        headers = list(rows[0].keys())

        # En-têtes
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Données
        for row_idx, row in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                val = row.get(header)
                if isinstance(val, (dict, list)):
                    val = json.dumps(val, ensure_ascii=False)
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Auto-largeur colonnes
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, min(len(str(cell.value)), 50))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2

        # Filtre auto
        ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"✓ Excel créé : {output_path}")


def upload_sftp(local_path: str, remote_filename: str):
    """Upload via expect + sftp sur OVH."""
    expect_script = f"""
set timeout 30
spawn sftp {OVH_USER}@{OVH_HOST}
expect "password:"
send "{OVH_PASS}\\r"
expect "sftp>"
send "mkdir {OVH_REMOTE_DIR}\\r"
expect "sftp>"
send "cd {OVH_REMOTE_DIR}\\r"
expect "sftp>"
send "put {local_path} {remote_filename}\\r"
expect "sftp>"
send "bye\\r"
expect eof
"""
    result = subprocess.run(
        ["expect", "-c", expect_script],
        capture_output=True, text=True, timeout=60
    )
    if result.returncode == 0:
        print(f"✓ Upload SFTP réussi : {OVH_REMOTE_DIR}/{remote_filename}")
    else:
        print(f"✗ Erreur SFTP : {result.stderr}")
        print(result.stdout)


def main():
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d_%Hh%M")
    filename = f"internlog_backup_{date_str}.xlsx"

    print(f"=== Backup InternLog — {now.strftime('%d/%m/%Y %H:%M')} ===\n")

    # 1. Fetch toutes les tables
    tables_data = {}
    for table in TABLES:
        print(f"  Téléchargement : {table}...", end=" ")
        rows = fetch_table(table)
        tables_data[table] = rows
        print(f"{len(rows)} lignes")

    # 2. Créer l'Excel
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    create_excel(tables_data, tmp_path)

    # Stats
    total_rows = sum(len(rows) for rows in tables_data.values())
    total_tables = len([t for t in tables_data.values() if t])
    print(f"\n📊 {total_tables} tables, {total_rows} lignes au total")

    # 3. Upload SFTP
    print(f"\n📤 Upload vers OVH ({OVH_HOST})...")
    upload_sftp(tmp_path, filename)

    # 4. Copie locale aussi
    local_backup_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "backups")
    os.makedirs(local_backup_dir, exist_ok=True)
    local_copy = os.path.join(local_backup_dir, filename)
    import shutil
    shutil.copy2(tmp_path, local_copy)
    print(f"✓ Copie locale : {local_copy}")

    # Cleanup
    os.unlink(tmp_path)
    print(f"\n=== Backup terminé ===")


if __name__ == "__main__":
    main()
